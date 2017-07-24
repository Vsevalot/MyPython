
#!/usr/bin/python

def ground_issues_detector(string):
    if (string == None): # if there is not any comment in a cell
        return 0
    for i in range(len(string)):
        if(string[i]=="O" or string[i]=="О" or string[i]=="O" or string[i]=="О"): # if found any mention about O1 or O2 electrode - issue found
            if(string[i+1]=="1" or string[i+1]=="2"):
                return 1
    return 0

def placebo_detector(string):
    if (string == "плацебо"):
        return 1
    return 0

from openpyxl import load_workbook

wb = load_workbook('Protocol1.xlsx')
sheet_1 = wb.get_sheet_by_name('Лист1')

time_phases=[]
column=[]
for i in range (1, 11):
    if(2==i):# date of experiment (do not need this data, so skip it)
        continue
    for k in range(2, sheet_1.max_row):
        column.append(sheet_1.cell(row=k, column=i).value)
    time_phases.append(column)
    column=[]

# finding correct records
correct_lines=[]
for i in range(0, len(time_phases[8])):
    if(not placebo_detector(time_phases[7][i]) and time_phases[7][i]!=None): # not a placebo patient or empty line
        if(not ground_issues_detector(time_phases[8][i])): # no any issues with O1 or O2 electrodes
            correct_lines.append(i)


# transpose
a=[]
b=[]
for i in range(0, len(time_phases[0])):
    for k in range(0, len(time_phases)):
        b.append(time_phases[k][i])
    a.append(b)
    b=[]

c=[]
for i in range(min(correct_lines), max(correct_lines)):
    if any([k==i for k in correct_lines]):
        c.append(a[i])


# convert time to second cuts
for i in range(len(c)):
    for k in range(len(c[i])):
        if(k>0 and k<7):
            try:
                c[i][k] = c[i][k].hour*3600 + c[i][k].minute*60
            except AttributeError:
                hours=""
                minutes=""
                n=0
                for z in range(len(c[i][k])):
                    if(c[i][k][z]==":"):
                        n=z # place of ':'
                for z in range(n):
                    hours+=c[i][k][z]
                hours=int(hours)*3600
                for z in range (n+1, len(c[i][k])):
                    minutes+=c[i][k][z]
                minutes=int(minutes)*60
                c[i][k]=hours+minutes


final=[]
b=[]
for i in range(0,len(c)):
    for k in range(0,len(c[i])-1):
        if(0==k or k==len(c[i])-2):
            b.append(c[i][k])
        elif(k==len(c[i])-3):
            continue
        else:
            b.append((c[i][k]-c[i][1])*1000)
    final.append(b)
    b=[]

print(final)

csvFile = open("Work_protocol"+".csv",'w')
string=""
for i in range(len(final)):
    for k in range(len(final[i])):
        string+=str(final[i][k])
        if(k==len(final[i])-1):
            string+='\n'
        else:
            string+=';'
    #print(string)
    csvFile.write(string)
    string=""

csvFile.close()



